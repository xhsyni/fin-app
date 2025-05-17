from django.shortcuts import render
from .models import Users, Income, Expense, File, TaxUser, Tax
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from rest_framework import viewsets
from rest_framework.response import Response
from .serializers import UserSerializer, IncomeSerializer, ExpenseSerializer, FileSerializer, TaxUserSerializer, TaxSerializer
from django.contrib.auth import authenticate
from django.contrib.auth.hashers import make_password, check_password
from rest_framework_simplejwt.tokens import RefreshToken
import os
from openai import OpenAI
from backend import settings


# Create your views here.
@api_view(['GET','POST'])
def userLogin(request):
    emailInput = request.data.get('email')
    passwordInput = request.data.get('password')
    print(make_password("abcd"))
    
    try:
        user = Users.objects.get(email=emailInput)
        if check_password(passwordInput, user.password):
            refresh = RefreshToken.for_user(user)

            return Response({
                "message": "Login successful",
                "user": {
                    "name": user.name,
                    "email": user.email,
                    "userid": user.userid,
                    "job":user.job,
                    "phone":user.phone,
                    "birth_date":user.birth_date,
                    "marital_status":user.marital_status,
                    "ethnicity":user.ethnicity,
                    "gender":user.gender,
                    "nationality":user.nationality
                },
                "token": {
                    "refresh": str(refresh),
                    "access": str(refresh.access_token)
                }
            }, status=status.HTTP_200_OK)
        else:
            return Response(
                {"error": "Invalid credentials"}, 
                status=status.HTTP_401_UNAUTHORIZED
            )
    except Users.DoesNotExist:
        return Response(
            {"error": "User not found"}, 
            status=status.HTTP_404_NOT_FOUND
        )

@api_view(['GET','POST'])
def signupUser(request):
    nameInput = request.data.get('name')
    emailInput = request.data.get('email')
    passwordInput = request.data.get('password')

    hashedPassword = make_password(passwordInput)
    Users.objects.create(
        name = nameInput,
        email = emailInput,
        password=hashedPassword,
    )
    return Response({
        "message": "Register successful"
    }, status=status.HTTP_200_OK)


@api_view(['GET','POST'])
@permission_classes([IsAuthenticated])
def updateUser(request):
    userIns = request.user
    nameInput = request.data.get('name')
    emailInput = request.data.get('email')
    phoneInput = request.data.get('phone')
    jobInput = request.data.get('job')
    birth_dateInput = request.data.get('birth_date') if request.data.get('birth_date') else None
    marital_statusInput = request.data.get('marital_status') if request.data.get('marital_status') else False
    ethnicityInput = request.data.get('ethnicity') if request.data.get('ethnicity') else None
    genderInput = request.data.get('gender') if request.data.get('gender') else None
    nationalityInput = request.data.get('nationality')if request.data.get('nationality') else None


    Users.objects.update(name=nameInput,phone=phoneInput,job=jobInput,birth_date=birth_dateInput,marital_status=marital_statusInput,ethnicity=ethnicityInput,gender=genderInput,nationality=nationalityInput)

    return Response({
        "message":"User information update successfully."
    }, status=status.HTTP_200_OK)


@api_view(['GET','POST'])
@permission_classes([IsAuthenticated])
def viewAllTax(request):
    userIns = request.user
    taxIns = TaxUser.objects.filter(userid=userIns).order_by('-taxuserid')
    
    totalIncomeData = {}
    taxUserData = []
    for tax in taxIns:
        year = tax.tax_year
        income = tax.tax_income
        amount = tax.amount 
        taxid = tax.taxuserid
        deductionExpense = tax.deduction
        taxUserData.append({
            "id":taxid,
            "year": year,
            "deduction": deductionExpense,
            "taxable_income": income,
            "income_tax": amount,
        })
        if year not in totalIncomeData:
            totalIncomeData[year] = {
                "taxable_income": 0,
                "income_tax": 0,
                "deduction":0
            }
        
        totalIncomeData[year]["taxable_income"] += income
        totalIncomeData[year]["income_tax"] += amount
        totalIncomeData[year]["deduction"] += deductionExpense
    
    return Response({
        "tax": taxUserData,
        "totalTax": totalIncomeData,
    }, status=status.HTTP_200_OK)

messages = [
    {
        "role": "system",
        "content": """You are a tax assistant helping freelancers with Malaysia tax rules. 
        Your job is to help users track income and expenses, explain deductible expenses, clarify tax rules, and answer tax-related questions clearly and politely. 
        If the user gives incomplete information, ask follow-up questions to get the details you need. """,
    }
]

def get_response(messages):
    client = OpenAI(
        api_key="sk-41805a630687481eb9e57feb30f512bc",
        base_url="https://dashscope-intl.aliyuncs.com/compatible-mode/v1",
    )
    completion = client.chat.completions.create(model="qwen-max", messages=messages)
    return completion

@api_view(['GET','POST'])
@permission_classes([IsAuthenticated])
def chatWithAI(request):
    userIns = request.user
    userInput = request.data.get('messages')
    
    last_message = userInput[-1]
    user_content = last_message.get('text', '')
    
    messages.append({"role": "user", "content": user_content})
    
    assistant_output = get_response(messages).choices[0].message.content
    messages.append({"role": "assistant", "content": assistant_output})

    return Response({
        "messageQwen": assistant_output,
    }, status=status.HTTP_200_OK)

taxTable = {
    "year":[
        {"category": "A", "min": 0, "max": 5000, "rate": 0, "fixed_tax": 0},
        {"category": "B", "min": 5001, "max": 20000, "rate": 1, "fixed_tax": 0},
        {"category": "C", "min": 20001, "max": 35000, "rate": 3, "fixed_tax": 150},
        {"category": "D", "min": 35001, "max": 50000, "rate": 6, "fixed_tax": 600},
        {"category": "E", "min": 50001, "max": 70000, "rate": 11, "fixed_tax": 1500},
        {"category": "F", "min": 70001, "max": 100000, "rate": 19, "fixed_tax": 3700},
        {"category": "G", "min": 100001, "max": 400000, "rate": 25, "fixed_tax": 9400},
        {"category": "H", "min": 400001, "max": 600000, "rate": 26, "fixed_tax": 84400},
        {"category": "I", "min": 600001, "max": 2000000, "rate": 28, "fixed_tax": 136400},
        {"category": "J", "min": 2000001, "max": float('inf'), "rate": 30, "fixed_tax": 528400},
    ]
}

def calculateTax(income):
    for yr in taxTable:
        yearTax = taxTable[yr]
        for rows in yearTax:
            if income >= rows["min"]:
                taxableIncome = income - (rows["min"]-1)
                tax = rows["fixed_tax"] + taxableIncome*(rows["rate"]/100)
                category = rows["category"]
            else:
                break
    if taxableIncome <= 35000:
        tax -= 400
    
    return tax, taxableIncome, category

@api_view(['GET','POST'])
@permission_classes([IsAuthenticated])
def addTax(request):
    userIns = request.user
    taxYear = str(request.data.get('year'))

    incomeIns = Income.objects.filter(userid=userIns)
    acceptableIncomeCategory = {"employment","business","rental","dividend","interest","royalties","pensions","foreign"}
    totalIncome = 0
    incomeIdList = []
    for income in incomeIns:
        incomeid = income.incomeid
        year = str(income.date).split("-")[0]
        incomeAmount = float(income.amount)
        incomeCategory = income.category        
        if taxYear == year:
            if income.typeInc.lower() == "gain":
                totalIncome += incomeAmount
            else:
                totalIncome -= incomeAmount
            incomeIdList.append(incomeid)

    acceptableExpenseCategory = {"self", "medical","lifestyle","education","child expenses","parenthood","insurance","family","disability","donation"}
    expenseIns = Expense.objects.filter(userid=userIns)
    totalExpense = 0
    expenseIdList = []
    for expense in expenseIns:
        expenseid = expense.expenseid
        year = str(expense.date).split("-")[0]
        expenseAmount = float(expense.amount)
        expenseCategory = expense.category
        if taxYear == year and expenseCategory.lower() in acceptableExpenseCategory:
            expenseIdList.append(expenseid)
            totalExpense += expenseAmount
    
    income = totalIncome - totalExpense
    tax, taxable_income, category = calculateTax(income)

    taxUserIns = TaxUser.objects.create(
        userid = userIns,
        tax_year = taxYear,
        tax_income=income,
        deduction=totalExpense,
        amount=tax
    )

    for income in incomeIdList:
        incomeToUpdate = Income.objects.filter(incomeid=income)
        incomeToUpdate.update(taxuserid=taxUserIns)
    for expense in expenseIdList:
        expenseToUpdate = Expense.objects.filter(expenseid=expense)
        expenseToUpdate.update(taxuserid=taxUserIns)

    return Response({
        "message": "Tax added successfully!",
    }, status=status.HTTP_200_OK)


@api_view(['GET','POST'])
@permission_classes([IsAuthenticated])
def updateTax(request):
    userIns = request.user
    taxYear = str(request.data.get('year'))

    incomeIns = Income.objects.filter(userid=userIns)
    acceptableIncomeCategory = {"employment","business","rental","dividend","interest","royalties","pensions","foreign"}
    totalIncome = 0
    incomeIdList = []
    for income in incomeIns :
        incomeid = income.incomeid
        year = str(income.date).split("-")[0]
        incomeAmount = float(income.amount)
        incomeCategory = income.category        
        if taxYear == year and incomeCategory.lower() in acceptableIncomeCategory:
            if income.typeInc.lower() == "gain":
                totalIncome += incomeAmount
            else:
                totalIncome -= incomeAmount
            incomeIdList.append(incomeid)

    acceptableExpenseCategory = {"self", "medical","lifestyle","education","child expenses","parenthood","insurance","family","disability","donation"}
    expenseIns = Expense.objects.filter(userid=userIns)
    totalExpense = 0
    expenseIdList = []
    for expense in expenseIns:
        expenseid = expense.expenseid
        year = str(expense.date).split("-")[0]
        expenseAmount = float(expense.amount)
        expenseCategory = expense.category
        if taxYear == year and expenseCategory.lower() in acceptableExpenseCategory:
            expenseIdList.append(expenseid)
            totalExpense += expenseAmount
    
    income = totalIncome - totalExpense
    tax, taxable_income, category = calculateTax(income)

    taxUserIns = TaxUser.objects.update(
        userid = userIns,
        tax_year = taxYear,
        tax_income=income,
        deduction=totalExpense,
        amount=tax
    )

    for income in incomeIdList:
        incomeToUpdate = Income.objects.filter(incomeid=income)
        incomeToUpdate.update(taxuserid=taxUserIns)
    for expense in expenseIdList:
        expenseToUpdate = Expense.objects.filter(expenseid=expense)
        expenseToUpdate.update(taxuserid=taxUserIns)

    return Response({
        "message": "Tax added successfully!",
    }, status=status.HTTP_200_OK)


from openpyxl import Workbook
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def exportReport(request):
    user = request.user
    tax_year = request.data.get('year')

    file_name = f"report_{tax_year or 'all'}.xlsx"
    dir_path = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(dir_path, exist_ok=True)
    file_path = os.path.join(dir_path, file_name)

    wb = Workbook()

    # Income sheet
    ws1 = wb.active
    ws1.title = "Income"
    ws1.append(["Date", "Category", "Source", "Amount", "Job", "Type"])
    incomes = Income.objects.filter(userid=user)
    for inc in incomes:
        ws1.append([
            inc.date.strftime('%Y-%m-%d'),
            inc.category,
            inc.source,
            float(inc.amount),
            inc.job or '',
            inc.typeInc or ''
        ])

    # Expense sheet
    ws2 = wb.create_sheet(title="Expense")
    ws2.append(["Date", "Category", "Sub Category", "Amount", "Name", "Deductible"])
    expenses = Expense.objects.filter(userid=user)
    for exp in expenses:
        ws2.append([
            exp.date.strftime('%Y-%m-%d'),
            exp.category,
            exp.sub_category,
            float(exp.amount),
            exp.name,
            "Yes" if exp.deductible else "No"
        ])

    # Expense sheet
    ws3 = wb.create_sheet(title="Tax")
    ws3.append(["tax_year", "tax_income", "deduction", "amount"])
    taxUser = TaxUser.objects.filter(userid=user)
    for tax in taxUser:
        ws3.append([
            tax.tax_year,
            tax.tax_income,
            tax.deduction,
            tax.amount,
        ])

    wb.save(file_path)

    file_url = os.path.join(settings.MEDIA_URL, 'reports', file_name).replace('\\', '/')
    return Response({"excel_url": file_url})


class UserViewSet(viewsets.ModelViewSet):
    queryset = Users.objects.all()
    serializer_class = UserSerializer

class IncomeViewSet(viewsets.ModelViewSet):
    queryset = Income.objects.all()
    serializer_class = IncomeSerializer

class ExpenseViewSet(viewsets.ModelViewSet):
    queryset = Expense.objects.all()
    serializer_class = ExpenseSerializer

class FileViewSet(viewsets.ModelViewSet):
    queryset = File.objects.all()
    serializer_class = FileSerializer

class TaxUserViewSet(viewsets.ModelViewSet):
    queryset = TaxUser.objects.all()
    serializer_class = TaxUserSerializer

class TaxViewSet(viewsets.ModelViewSet):
    queryset = Tax.objects.all()
    serializer_class = TaxSerializer

